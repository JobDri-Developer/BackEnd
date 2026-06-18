#!/usr/bin/env python3
"""Import corpus data from xlsx into Postgres.

Usage:
    python scripts/import_corpus.py /path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    import psycopg
except ImportError:  # pragma: no cover - fallback for older environments
    psycopg = None

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:  # pragma: no cover - optional fallback
    psycopg2 = None
    RealDictCursor = None


JD_SHEET_NAME = "jd_embed_corpus"
QUESTION_SHEET_NAME = "question_embed_corpus"


@dataclass
class ImportStats:
    created_companies: int = 0
    created_job_postings: int = 0
    updated_job_postings: int = 0
    created_questions: int = 0
    updated_questions: int = 0
    matched_classifications: int = 0
    unmatched_classifications: int = 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import corpus xlsx into Postgres.")
    parser.add_argument("xlsx_path", help="Path to xlsx file")
    parser.add_argument(
        "--env-file",
        default=".env",
        help="Environment file to read DB_URL/DB_USERNAME/DB_PASSWORD from (default: .env)",
    )
    return parser.parse_args()


def load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


def jdbc_to_postgres_dsn(jdbc_url: str) -> str:
    prefix = "jdbc:"
    return jdbc_url[len(prefix):] if jdbc_url.startswith(prefix) else jdbc_url


def connect():
    db_url = os.environ.get("DB_URL")
    db_user = os.environ.get("DB_USERNAME")
    db_password = os.environ.get("DB_PASSWORD")

    if not db_url or not db_user:
        raise SystemExit("DB_URL and DB_USERNAME must be set in environment or env file.")

    dsn = jdbc_to_postgres_dsn(db_url)
    if psycopg is not None:
        return psycopg.connect(dsn, user=db_user, password=db_password)
    if psycopg2 is not None:
        return psycopg2.connect(dsn, user=db_user, password=db_password, cursor_factory=RealDictCursor)
    raise SystemExit("Install psycopg or psycopg2-binary before running this script.")


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_string(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def to_int(value: Any) -> int | None:
    text = to_string(value)
    if text is None:
        return None
    return int(float(text))


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = to_string(value)
    if text is None:
        return False
    lowered = text.lower()
    if lowered in {"true", "t", "1", "y", "yes"}:
        return True
    if lowered in {"false", "f", "0", "n", "no"}:
        return False
    raise ValueError(f"Unsupported boolean value: {value}")


def build_header_map(sheet) -> dict[str, int]:
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {}
    return {str(cell).strip(): idx for idx, cell in enumerate(header_row) if cell is not None}


def get_cell(row: tuple[Any, ...], header_map: dict[str, int], key: str) -> Any:
    idx = header_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def require_headers(header_map: dict[str, int], *headers: str) -> None:
    missing = [header for header in headers if header not in header_map]
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")


def fetch_one(cur, query: str, params: tuple[Any, ...]):
    cur.execute(query, params)
    row = cur.fetchone()
    if row is None:
        return None
    if isinstance(row, dict):
        return row
    if hasattr(row, "_mapping"):
        return dict(row._mapping)
    desc = cur.description
    return {desc[i].name if hasattr(desc[i], "name") else desc[i][0]: row[i] for i in range(len(row))}


def resolve_company_id(cur, cache: dict[str, int], stats: ImportStats, company_name: str | None) -> int | None:
    if not company_name:
        return None
    cached = cache.get(company_name)
    if cached is not None:
        return cached

    row = fetch_one(cur, "select id from companies where name = %s", (company_name,))
    if row is not None:
        company_id = int(row["id"])
        cache[company_name] = company_id
        return company_id

    row = fetch_one(
        cur,
        """
        insert into companies (name, size, created_at, updated_at)
        values (%s, %s, now(), now())
        returning id
        """,
        (company_name, None),
    )
    company_id = int(row["id"])
    cache[company_name] = company_id
    stats.created_companies += 1
    return company_id


def resolve_detail_classification_id(cur, cache: dict[tuple[str | None, str | None, str | None], int | None], stats: ImportStats, job_group_l1: str | None, job_family_l2: str | None, role_l3: str | None) -> int | None:
    key = (normalize_text(job_group_l1), normalize_text(job_family_l2), normalize_text(role_l3))
    if key in cache:
        classification_id = cache[key]
        if classification_id is not None:
            stats.matched_classifications += 1
        elif key[2]:
            stats.unmatched_classifications += 1
        return classification_id

    normalized_job_group, normalized_job_family, normalized_role = key
    if not normalized_role:
        cache[key] = None
        return None

    row = fetch_one(
        cur,
        """
        select ccm.detail_classification_id as id
        from corpus_classification_mappings ccm
        join detail_classifications dcm on dcm.id = ccm.detail_classification_id
        where ccm.source_job_group_l1 = %s
          and ccm.source_job_family_l2 = %s
          and ccm.source_role_l3 = %s
        """,
        (normalized_job_group, normalized_job_family, normalized_role),
    )
    if row is None and normalized_job_group and normalized_job_family:
        row = fetch_one(
            cur,
            """
            select dc.id
            from detail_classifications dc
            join middle_classifications mc on dc.middle_classification_id = mc.id
            join classifications c on mc.classification_id = c.id
            where lower(c.big_name) = lower(%s)
              and lower(mc.middle_name) = lower(%s)
              and lower(dc.detail_name) = lower(%s)
            """,
            (normalized_job_group, normalized_job_family, normalized_role),
        )
    if row is None:
        row = fetch_one(
            cur,
            """
            select dc.id
            from detail_classifications dc
            where lower(dc.detail_name) = lower(%s)
              and 1 = (
                  select count(*)
                  from detail_classifications dc2
                  where lower(dc2.detail_name) = lower(%s)
              )
            """,
            (normalized_role, normalized_role),
        )

    classification_id = int(row["id"]) if row is not None else None
    cache[key] = classification_id
    if classification_id is not None:
        stats.matched_classifications += 1
    else:
        stats.unmatched_classifications += 1
    return classification_id


def upsert_job_posting(cur, payload: dict[str, Any]) -> bool:
    row = fetch_one(
        cur,
        """
        insert into mock_job_posting_corpus (
            source_analysis_id,
            company_id,
            detail_classification_id,
            company_name,
            industry,
            job_group_l1,
            job_family_l2,
            role_l3,
            skills,
            responsibilities,
            requirements,
            preferred,
            embedding_text,
            is_valid_for_embedding,
            invalid_reason,
            created_at,
            updated_at
        )
        values (
            %(source_analysis_id)s,
            %(company_id)s,
            %(detail_classification_id)s,
            %(company_name)s,
            %(industry)s,
            %(job_group_l1)s,
            %(job_family_l2)s,
            %(role_l3)s,
            %(skills)s,
            %(responsibilities)s,
            %(requirements)s,
            %(preferred)s,
            %(embedding_text)s,
            %(is_valid_for_embedding)s,
            %(invalid_reason)s,
            now(),
            now()
        )
        on conflict (source_analysis_id) do update
        set company_id = excluded.company_id,
            detail_classification_id = excluded.detail_classification_id,
            company_name = excluded.company_name,
            industry = excluded.industry,
            job_group_l1 = excluded.job_group_l1,
            job_family_l2 = excluded.job_family_l2,
            role_l3 = excluded.role_l3,
            skills = excluded.skills,
            responsibilities = excluded.responsibilities,
            requirements = excluded.requirements,
            preferred = excluded.preferred,
            embedding_text = excluded.embedding_text,
            is_valid_for_embedding = excluded.is_valid_for_embedding,
            invalid_reason = excluded.invalid_reason,
            updated_at = now()
        returning (xmax = 0) as inserted
        """,
        payload,
    )
    return bool(row["inserted"])


def upsert_question(cur, payload: dict[str, Any]) -> bool:
    row = fetch_one(
        cur,
        """
        insert into mock_question_corpus (
            source_question_id,
            source_analysis_id,
            company_id,
            detail_classification_id,
            company_name,
            job_group_l1,
            job_family_l2,
            role_l3,
            source,
            question_type,
            char_limit,
            question_text,
            embedding_text,
            is_valid_for_embedding,
            created_at,
            updated_at
        )
        values (
            %(source_question_id)s,
            %(source_analysis_id)s,
            %(company_id)s,
            %(detail_classification_id)s,
            %(company_name)s,
            %(job_group_l1)s,
            %(job_family_l2)s,
            %(role_l3)s,
            %(source)s,
            %(question_type)s,
            %(char_limit)s,
            %(question_text)s,
            %(embedding_text)s,
            %(is_valid_for_embedding)s,
            now(),
            now()
        )
        on conflict (source_question_id) do update
        set source_analysis_id = excluded.source_analysis_id,
            company_id = excluded.company_id,
            detail_classification_id = excluded.detail_classification_id,
            company_name = excluded.company_name,
            job_group_l1 = excluded.job_group_l1,
            job_family_l2 = excluded.job_family_l2,
            role_l3 = excluded.role_l3,
            source = excluded.source,
            question_type = excluded.question_type,
            char_limit = excluded.char_limit,
            question_text = excluded.question_text,
            embedding_text = excluded.embedding_text,
            is_valid_for_embedding = excluded.is_valid_for_embedding,
            updated_at = now()
        returning (xmax = 0) as inserted
        """,
        payload,
    )
    return bool(row["inserted"])


def import_job_posting_sheet(cur, sheet, stats: ImportStats, company_cache: dict[str, int], classification_cache: dict[tuple[str | None, str | None, str | None], int | None]) -> None:
    if sheet is None:
        return

    header_map = build_header_map(sheet)
    if not header_map:
        return
    require_headers(
        header_map,
        "analysis_id", "company_name", "job_group_l1", "job_family_l2", "role_l3",
        "skills", "responsibilities", "requirements", "preferred", "embedding_text", "is_valid_for_embedding",
    )

    for row in sheet.iter_rows(min_row=2, values_only=True):
        source_analysis_id = to_string(get_cell(row, header_map, "analysis_id"))
        if not source_analysis_id:
            continue

        company_name = to_string(get_cell(row, header_map, "company_name"))
        company_id = resolve_company_id(cur, company_cache, stats, company_name)
        detail_classification_id = resolve_detail_classification_id(
            cur,
            classification_cache,
            stats,
            to_string(get_cell(row, header_map, "job_group_l1")),
            to_string(get_cell(row, header_map, "job_family_l2")),
            to_string(get_cell(row, header_map, "role_l3")),
        )

        inserted = upsert_job_posting(
            cur,
            {
                "source_analysis_id": source_analysis_id,
                "company_id": company_id,
                "detail_classification_id": detail_classification_id,
                "company_name": company_name,
                "industry": to_string(get_cell(row, header_map, "industry")),
                "job_group_l1": to_string(get_cell(row, header_map, "job_group_l1")),
                "job_family_l2": to_string(get_cell(row, header_map, "job_family_l2")),
                "role_l3": to_string(get_cell(row, header_map, "role_l3")),
                "skills": to_string(get_cell(row, header_map, "skills")),
                "responsibilities": to_string(get_cell(row, header_map, "responsibilities")),
                "requirements": to_string(get_cell(row, header_map, "requirements")),
                "preferred": to_string(get_cell(row, header_map, "preferred")),
                "embedding_text": to_string(get_cell(row, header_map, "embedding_text")),
                "is_valid_for_embedding": to_bool(get_cell(row, header_map, "is_valid_for_embedding")),
                "invalid_reason": to_string(get_cell(row, header_map, "invalid_reason")),
            },
        )
        if inserted:
            stats.created_job_postings += 1
        else:
            stats.updated_job_postings += 1


def import_question_sheet(cur, sheet, stats: ImportStats, company_cache: dict[str, int], classification_cache: dict[tuple[str | None, str | None, str | None], int | None]) -> None:
    if sheet is None:
        return

    header_map = build_header_map(sheet)
    if not header_map:
        return
    require_headers(
        header_map,
        "question_id", "analysis_id", "company_name", "job_group_l1", "job_family_l2", "role_l3",
        "source", "question_text", "embedding_text", "is_valid_for_embedding",
    )

    for row in sheet.iter_rows(min_row=2, values_only=True):
        source_question_id = to_string(get_cell(row, header_map, "question_id"))
        if not source_question_id:
            continue

        company_name = to_string(get_cell(row, header_map, "company_name"))
        company_id = resolve_company_id(cur, company_cache, stats, company_name)
        detail_classification_id = resolve_detail_classification_id(
            cur,
            classification_cache,
            stats,
            to_string(get_cell(row, header_map, "job_group_l1")),
            to_string(get_cell(row, header_map, "job_family_l2")),
            to_string(get_cell(row, header_map, "role_l3")),
        )

        inserted = upsert_question(
            cur,
            {
                "source_question_id": source_question_id,
                "source_analysis_id": to_string(get_cell(row, header_map, "analysis_id")),
                "company_id": company_id,
                "detail_classification_id": detail_classification_id,
                "company_name": company_name,
                "job_group_l1": to_string(get_cell(row, header_map, "job_group_l1")),
                "job_family_l2": to_string(get_cell(row, header_map, "job_family_l2")),
                "role_l3": to_string(get_cell(row, header_map, "role_l3")),
                "source": to_string(get_cell(row, header_map, "source")),
                "question_type": to_string(get_cell(row, header_map, "question_type")),
                "char_limit": to_int(get_cell(row, header_map, "char_limit")),
                "question_text": to_string(get_cell(row, header_map, "question_text")),
                "embedding_text": to_string(get_cell(row, header_map, "embedding_text")),
                "is_valid_for_embedding": to_bool(get_cell(row, header_map, "is_valid_for_embedding")),
            },
        )
        if inserted:
            stats.created_questions += 1
        else:
            stats.updated_questions += 1


def main() -> int:
    args = parse_args()
    xlsx_path = Path(args.xlsx_path).expanduser().resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"XLSX file not found: {xlsx_path}")

    load_env_file(Path(args.env_file))
    workbook = load_workbook(filename=xlsx_path, data_only=True)

    stats = ImportStats()
    company_cache: dict[str, int] = {}
    classification_cache: dict[tuple[str | None, str | None, str | None], int | None] = {}

    conn = connect()
    try:
        with conn:
            with conn.cursor() as cur:
                import_job_posting_sheet(
                    cur,
                    workbook[JD_SHEET_NAME] if JD_SHEET_NAME in workbook.sheetnames else None,
                    stats,
                    company_cache,
                    classification_cache,
                )
                import_question_sheet(
                    cur,
                    workbook[QUESTION_SHEET_NAME] if QUESTION_SHEET_NAME in workbook.sheetnames else None,
                    stats,
                    company_cache,
                    classification_cache,
                )
    finally:
        workbook.close()
        conn.close()

    print("Import completed")
    print(f"createdCompanies={stats.created_companies}")
    print(f"createdJobPostings={stats.created_job_postings}")
    print(f"updatedJobPostings={stats.updated_job_postings}")
    print(f"createdQuestions={stats.created_questions}")
    print(f"updatedQuestions={stats.updated_questions}")
    print(f"matchedClassifications={stats.matched_classifications}")
    print(f"unmatchedClassifications={stats.unmatched_classifications}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - CLI error surface
        print(f"Import failed: {exc}", file=sys.stderr)
        raise
